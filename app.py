# Projeto: Extrator de PDF -> Excel (usando uma planilha modelo existente)
# Estrutura sugerida:
# projeto_extrator/
# ├── app.py
# ├── planilha_teste.xlsx         # <-- sua planilha modelo (colunas/aba já configuradas)
# ├── templates/
# │   └── index.html
# └── uploads/                    # arquivos temporários (pdfs e xlsx gerados)

"""
Este arquivo é o backend principal (Flask).
Comportamento:
 - Recebe upload de PDF
 - Extrai dados com pdfplumber (mesma lógica que você já tinha)
 - Copia a planilha modelo (planilha_teste.xlsx) para um nome temporário
 - Abre a cópia com openpyxl e preenche a aba especificada
 - Retorna o arquivo .xlsx gerado como download

Instalação:
pip install flask pdfplumber openpyxl Werkzeug

Como rodar:
python app.py
Abrir: http://127.0.0.1:5000

Coloque sua planilha modelo em planilha_teste.xlsx (mesmo diretório do app.py). A aba alvo deve existir (nome padrão: 'Omie_Contas_Pagar').
"""

import os
import shutil
import time
from datetime import datetime
from flask import Flask, render_template, request, send_file, redirect, url_for, flash
from werkzeug.utils import secure_filename
import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# CONFIGURAÇÕES
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
ALLOWED_EXTENSIONS = {"pdf"}
TEMPLATE_XLSX = os.path.join(os.path.dirname(__file__), "planilha_teste.xlsx")
ABA_DESTINO = "Omie_Contas_Pagar"  # altere se necessário
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.secret_key = "troque_essa_chave_para_uma_secreta_em_producao"


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def extrair_dados_do_pdf(pdf_path):
    registros = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tabela = page.extract_table()
            if tabela:
                for linha in tabela[1:]:  # Ignora cabeçalho
                    # protege contra linhas curtas
                    if not linha or len(linha) < 2:
                        continue

                    nome = (linha[1] or "").strip()
                    if not nome or nome.upper() == "NOME DO PROFISSIONAL":
                        continue

                    valor_total = 0.0
                    chave_pix = ""

                    # Ajuste de índices: adapte conforme sua tabela real
                    # No seu script original você usava linha[14] para 'TOTAL A PAGAR' e linha[3] para PIX
                    # Aqui mantivemos essa lógica, mas valide com seus PDFs reais.
                    try:
                        if len(linha) > 14 and linha[14]:
                            raw = linha[14].replace("R$", "").replace(".", "").replace(",", ".").strip()
                            valor_total = float(raw) if raw not in ("", None) else 0.0
                    except Exception:
                        valor_total = 0.0

                    if len(linha) > 3 and linha[3]:
                        chave_pix = linha[3].strip()

                    registros.append({
                        'Fornecedor': nome,
                        'Categoria': 'Prestação de Serviços Médicos',
                        'Conta Corrente': 'Omie.CASH',
                        'Valor da Conta': valor_total,
                        'Projeto': 'Projeto Padrão',
                        'Data de Emissão': datetime.now().strftime('%d/%m/%Y'),
                        'Data de Registro': datetime.now().strftime('%d/%m/%Y'),
                        'Data de Vencimento': '15/08/2025',
                        'Valor do Pagamento': valor_total,
                        'Data de Conciliação': '',
                        'Observações': '',
                        'Chave Pix': chave_pix,
                        'Departamento (100%)': 'Financeiro'
                    })
    return registros


def preencher_planilha_template(dados, caminho_template, aba_destino):
    # Cria cópia do template para não sobrescrever o original
    timestamp = int(time.time())
    resultado_nome = f"resultado_{timestamp}.xlsx"
    resultado_path = os.path.join(UPLOAD_FOLDER, resultado_nome)
    shutil.copyfile(caminho_template, resultado_path)

    wb = load_workbook(resultado_path)

    if aba_destino not in wb.sheetnames:
        raise ValueError(f"A aba '{aba_destino}' não existe na planilha modelo.")

    ws = wb[aba_destino]

    # Mapeamento de colunas (mesmo que você já usava)
    colunas_mapeamento = {
        'Fornecedor': 'B',
        'Categoria': 'C',
        'Conta Corrente': 'D',
        'Valor da Conta': 'E',
        'Projeto': 'G',
        'Data de Emissão': 'H',
        'Data de Registro': 'I',
        'Data de Vencimento': 'J',
        'Valor do Pagamento': 'M',
        'Data de Conciliação': 'Q',
        'Observações': 'R',
        'Chave Pix': 'AJ',
        'Departamento (100%)': 'AW'
    }

    # Descobre a primeira linha livre — assumimos que o cabeçalho ocupa a primeira linha.
    start_row = ws.max_row + 1

    # Se a planilha do template já tiver apenas cabeçalho e sem dados, max_row pode ser 1.
    # Preenchendo a partir de start_row
    for i, registro in enumerate(dados):
        linha = start_row + i
        for col_nome, col_letra in colunas_mapeamento.items():
            col_idx = column_index_from_string(col_letra)
            ws.cell(row=linha, column=col_idx, value=registro.get(col_nome, ""))

    wb.save(resultado_path)
    return resultado_path


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        if 'pdf_file' not in request.files:
            flash('Nenhum arquivo enviado')
            return redirect(request.url)

        file = request.files['pdf_file']
        if file.filename == '':
            flash('Nome de arquivo inválido')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            saved_pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(saved_pdf_path)

            # Extrair dados
            try:
                dados = extrair_dados_do_pdf(saved_pdf_path)
            except Exception as e:
                flash(f'Erro ao extrair dados do PDF: {e}')
                return redirect(request.url)

            if not dados:
                flash('Nenhum registro encontrado no PDF')
                return redirect(request.url)

            # Preencher planilha (usa o TEMPLATE_XLSX)
            try:
                resultado_path = preencher_planilha_template(dados, TEMPLATE_XLSX, ABA_DESTINO)
            except Exception as e:
                flash(f'Erro ao preencher a planilha modelo: {e}')
                return redirect(request.url)

            # Fornece o arquivo gerado para download
            return send_file(resultado_path, as_attachment=True)

        else:
            flash('Tipo de arquivo não permitido. Envie um PDF.')
            return redirect(request.url)

    # GET
    return render_template('index.html')


if __name__ == '__main__':
    # Verifica existência do template antes de iniciar
    if not os.path.exists(TEMPLATE_XLSX):
        print(f"Arquivo de template não encontrado: {TEMPLATE_XLSX}")
        print("Coloque sua planilha modelo com o nome 'planilha_teste.xlsx' no mesmo diretório do app.py")
    app.run(debug=True)
